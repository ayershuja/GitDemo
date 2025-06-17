import openpyxl
book=openpyxl.load_workbook("C:\\Users\\ayers\\Documents\\PythonDemo.xlsx")

sheet=book.active
cell=sheet.cell(row=1,column=2)
print(cell.value)

UpdateCell=sheet.cell(row=2,column=2).value="Ayer"
print(UpdateCell)


print(sheet['A1'].value)

print(sheet.max_row)

#print all value in excel sheet
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="Testcase2":
        for j in range(2,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)
